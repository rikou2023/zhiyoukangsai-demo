from __future__ import annotations

import json
import mimetypes
import os
import re
import sys
import uuid
from datetime import datetime, timezone
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Any
from urllib.parse import unquote, urlparse

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
WEBSITE_DIR = ROOT / "website"
LOGO_DIR = WEBSITE_DIR / "logo" / "logo"
DATABASE_DIR = ROOT / "database"
JSON_DB = DATABASE_DIR / "leads.json"
EXCEL_DB = DATABASE_DIR / "leads.xlsx"

ALLOWED_AI_MODELS = {
    "Kimi",
    "文心一言",
    "通义千问",
    "豆包",
    "智谱清言",
    "DeepSeek",
}

ALLOWED_PAIN_POINTS = {
    "A",  # AI 搜不到我们公司/产品
    "B",  # AI 总是优先推荐竞争对手
    "C",  # AI 对我们的回答有错误/负面信息
    "D",  # 想建立专属的企业 AI 知识库
}

PAIN_POINT_LABELS = {
    "A": "A · AI 搜不到我们公司/产品",
    "B": "B · AI 总是优先推荐竞争对手",
    "C": "C · AI 对我们的回答有错误/负面信息",
    "D": "D · 想建立专属的企业 AI 知识库",
}

HEADERS = [
    "提交时间",
    "线索ID",
    "公司名称",
    "姓名",
    "职位",
    "联系电话",
    "邮箱",
    "核心需求",
    "关注的 AI 大模型",
    "痛点选项",
    "来源页面",
    "IP",
    "User-Agent",
]


class AppError(Exception):
    def __init__(self, status: int, message: str) -> None:
        super().__init__(message)
        self.status = status
        self.message = message


def ensure_database() -> None:
    DATABASE_DIR.mkdir(parents=True, exist_ok=True)
    if not JSON_DB.exists():
        JSON_DB.write_text("[]", encoding="utf-8")
    records = load_records()
    write_excel(records)


def load_records() -> list[dict[str, Any]]:
    if not JSON_DB.exists():
        return []
    try:
        data = json.loads(JSON_DB.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        backup = JSON_DB.with_suffix(f".broken-{datetime.now().strftime('%Y%m%d%H%M%S')}.json")
        JSON_DB.replace(backup)
        JSON_DB.write_text("[]", encoding="utf-8")
        return []
    return data if isinstance(data, list) else []


def save_records(records: list[dict[str, Any]]) -> None:
    tmp = JSON_DB.with_suffix(".json.tmp")
    tmp.write_text(json.dumps(records, ensure_ascii=False, indent=2), encoding="utf-8")
    tmp.replace(JSON_DB)
    write_excel(records)


def write_excel(records: list[dict[str, Any]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "联系我们表单"

    title_fill = PatternFill("solid", fgColor="0B1220")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    title_font = Font(color="FFFFFF", bold=True, size=16)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
    title = ws.cell(row=1, column=1, value="智优康赛官网线索收集表")
    title.fill = title_fill
    title.font = title_font
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    for col, header in enumerate(HEADERS, 1):
        item = ws.cell(row=2, column=col, value=header)
        item.fill = header_fill
        item.font = header_font
        item.border = border
        item.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_index, record in enumerate(records, 3):
        # 新规范字段优先，旧字段名 fallback（不丢历史数据）
        pain_code = record.get("painPoint", "")
        pain_label = PAIN_POINT_LABELS.get(pain_code, pain_code)
        role_val = record.get("role") or record.get("position", "")
        message_val = record.get("message") or record.get("businessStatus", "")
        ai_focus_val = record.get("aiFocus") or record.get("primaryInterest") or []
        values = [
            record.get("createdAt", ""),
            record.get("id", ""),
            record.get("company", ""),
            record.get("name", ""),
            role_val,
            record.get("phone", ""),
            record.get("email", ""),
            message_val,
            "、".join(ai_focus_val) if isinstance(ai_focus_val, list) else str(ai_focus_val),
            pain_label,
            record.get("source", ""),
            record.get("ip", ""),
            record.get("userAgent", ""),
        ]
        for col, value in enumerate(values, 1):
            item = ws.cell(row=row_index, column=col, value=value)
            item.border = border
            item.alignment = Alignment(vertical="top", wrap_text=True)
            if row_index % 2 == 1:
                item.fill = PatternFill("solid", fgColor="F8FBFF")

    widths = [22, 18, 24, 12, 16, 18, 28, 48, 30, 36, 18, 18, 46]
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(HEADERS))}{max(len(records) + 2, 2)}"
    EXCEL_DB.parent.mkdir(parents=True, exist_ok=True)
    wb.save(EXCEL_DB)


def clean_text(value: Any, max_length: int) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    text = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", "", text)
    return text[:max_length]


def validate_contact(payload: dict[str, Any]) -> dict[str, Any]:
    """新规范字段：company / name / role / phone / email / message / ai_focus[] / pain_point"""
    if not isinstance(payload, dict):
        raise AppError(400, "请求格式不正确。")

    company = clean_text(payload.get("company"), 120)
    name = clean_text(payload.get("name"), 80)
    role = clean_text(payload.get("role"), 80)
    phone = clean_text(payload.get("phone"), 60)
    email = clean_text(payload.get("email"), 160)
    message = clean_text(payload.get("message"), 1200)
    source = clean_text(payload.get("source"), 200)

    # AI 大模型多选（数组）
    ai_focus_raw = payload.get("ai_focus", [])
    if isinstance(ai_focus_raw, str):
        ai_focus = [ai_focus_raw]
    elif isinstance(ai_focus_raw, list):
        ai_focus = [clean_text(item, 40) for item in ai_focus_raw]
    else:
        ai_focus = []
    ai_focus = [item for item in ai_focus if item in ALLOWED_AI_MODELS]

    # 痛点单选（A/B/C/D）
    pain_point = clean_text(payload.get("pain_point"), 4).upper()
    if pain_point not in ALLOWED_PAIN_POINTS:
        pain_point = ""

    missing = []
    if not company:    missing.append("公司名称")
    if not name:       missing.append("姓名")
    if not role:       missing.append("职位")
    if not phone:      missing.append("联系电话")
    if not email:      missing.append("邮箱")
    if not message:    missing.append("核心需求")
    if not ai_focus:   missing.append("最关注的 AI 大模型")
    if not pain_point: missing.append("品牌在 AI 时代的痛点")
    if missing:
        raise AppError(400, "请填写：" + "、".join(missing))
    if not re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", email):
        raise AppError(400, "邮箱格式不正确。")

    return {
        "company":    company,
        "name":       name,
        "role":       role,
        "phone":      phone,
        "email":      email,
        "message":    message,
        "aiFocus":    ai_focus,
        "painPoint":  pain_point,
        "source":     source or "website",
    }


def list_logos() -> list[dict[str, str]]:
    if not LOGO_DIR.exists():
        return []
    logos = []
    for path in sorted(LOGO_DIR.iterdir()):
        if path.is_file() and path.suffix.lower() in {".png", ".jpg", ".jpeg", ".webp", ".svg"}:
            rel = path.relative_to(WEBSITE_DIR).as_posix()
            logos.append({"name": path.stem, "url": rel})
    return logos


class Handler(SimpleHTTPRequestHandler):
    server_version = "ZhiyouKangsaiServer/1.0"

    def translate_path(self, path: str) -> str:
        parsed = urlparse(path)
        clean_path = unquote(parsed.path.lstrip("/"))
        if clean_path.startswith("api/") or clean_path == "":
            clean_path = "index.html" if clean_path == "" else clean_path
        resolved = (WEBSITE_DIR / clean_path).resolve()
        try:
            resolved.relative_to(WEBSITE_DIR.resolve())
        except ValueError:
            resolved = WEBSITE_DIR / "index.html"
        if resolved.is_dir():
            resolved = resolved / "index.html"
        return str(resolved)

    def end_headers(self) -> None:
        self.send_header("X-Content-Type-Options", "nosniff")
        self.send_header("Referrer-Policy", "strict-origin-when-cross-origin")
        self.send_header("Access-Control-Allow-Origin", self.headers.get("Origin", "*"))
        self.send_header("Access-Control-Allow-Methods", "GET, POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        super().end_headers()

    def do_OPTIONS(self) -> None:
        self.send_response(204)
        self.end_headers()

    def do_GET(self) -> None:
        parsed = urlparse(self.path)
        if parsed.path == "/health":
            self.write_json(200, {"ok": True, "service": "zhiyou-kangsai-website"})
            return
        if parsed.path == "/ready":
            self.write_json(200, {"ok": JSON_DB.exists() and EXCEL_DB.exists()})
            return
        if parsed.path == "/api/logos":
            self.write_json(200, {"ok": True, "logos": list_logos()})
            return
        return super().do_GET()

    def do_POST(self) -> None:
        parsed = urlparse(self.path)
        if parsed.path != "/api/contact":
            self.write_json(404, {"ok": False, "error": "接口不存在。"})
            return
        try:
            length = int(self.headers.get("Content-Length", "0"))
            if length > 1024 * 64:
                raise AppError(413, "提交内容过大。")
            raw = self.rfile.read(length).decode("utf-8")
            payload = json.loads(raw or "{}")
            data = validate_contact(payload)
            record = {
                "id": uuid.uuid4().hex[:12],
                "createdAt": datetime.now(timezone.utc).astimezone().strftime("%Y-%m-%d %H:%M:%S %z"),
                "ip": self.client_address[0],
                "userAgent": clean_text(self.headers.get("User-Agent"), 500),
                **data,
            }
            records = load_records()
            records.append(record)
            save_records(records)
            self.write_json(201, {"ok": True, "message": "您的需求已收到，会在3个工作日内给您反馈", "id": record["id"]})
        except AppError as error:
            self.write_json(error.status, {"ok": False, "error": error.message})
        except json.JSONDecodeError:
            self.write_json(400, {"ok": False, "error": "JSON格式不正确。"})
        except Exception as error:
            print(f"Unhandled error: {error}", file=sys.stderr)
            self.write_json(500, {"ok": False, "error": "服务器暂时不可用，请稍后重试。"})

    def write_json(self, status: int, data: dict[str, Any]) -> None:
        body = json.dumps(data, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def guess_type(self, path: str) -> str:
        if path.endswith(".js"):
            return "application/javascript"
        if path.endswith(".svg"):
            return "image/svg+xml"
        return mimetypes.guess_type(path)[0] or "application/octet-stream"


def main() -> None:
    ensure_database()
    host = os.environ.get("HOST", "127.0.0.1")
    port = int(os.environ.get("PORT", "8000"))
    httpd = ThreadingHTTPServer((host, port), Handler)
    print(f"智优康赛官网服务已启动: http://{host}:{port}")
    print(f"表单Excel数据库: {EXCEL_DB}")
    try:
        httpd.serve_forever()
    except KeyboardInterrupt:
        print("\n正在关闭服务")
    finally:
        httpd.server_close()


if __name__ == "__main__":
    main()
